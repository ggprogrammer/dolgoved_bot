import requests as req
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, Border, Side
from copy import copy
import os
from tg_sender import run_sender


url = "https://lk.mirea.ru/auth.php"

BASE_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'Dolgoved')
READFILE_DIR = os.path.join(BASE_DIR, 'READ_FILE')

def auth():
    session = req.Session()
    LOGIN = input("Введите логин: ")
    PASSWORD = input("Введите пароль: ")
    log_params = {
        "AUTH_FORM": "Y",
        "TYPE": "AUTH",
        "USER_LOGIN": LOGIN,
        "USER_PASSWORD": PASSWORD,
        "USER_REMEMBER": "Y"
    }

    r = session.post(url, log_params)
    r = session.get("https://lk.mirea.ru/learning/debt/")
    return r.text

with open("index.html", "w", encoding="utf-8") as f:
    f.write(auth())

def parser():
    with open("index.html", "r", encoding="utf-8") as f:
        responce = f
        soup = BeautifulSoup(responce, 'lxml')
        block = [i.text for i in soup.find_all("td")[::2]]
        res = []
        for el in block:
            d_name = el[:el.find("(")].strip()
            d_type = el[el.find("(")+1:].split(",")[0].strip()
            d_part = el[el.find("(")+1:].split(",")[1].strip()
            res.append([d_name, d_type, d_part])
        return res
    
def dop_ved_finder(dolgy):
    wb = openpyxl.load_workbook(README_FILE)
    dop_ved_dict = {}
    format_dict = {}
    for sheet in wb.worksheets:
        if sheet.title == "Справка":
            continue
        empty_counter = 0
        i = 0
        while empty_counter != 5:
            i += 1
            if sheet[f"A{i}"].value is None:
                empty_counter += 1
            elif sheet[f"A{i}"].value in dolgy:
                el = sheet[f"A{i}"].value
                dop_ved_dict.setdefault(sheet.title, []).append([el[:el.find("(")].strip(), el[el.find("("):].strip(), sheet[f"B{i}"].value.capitalize()])
                format_dict.setdefault(sheet.title, []).append(i)
    return dop_ved_dict, format_dict

def main_info_finder(dolgy):
    wb = openpyxl.load_workbook(README_FILE)
    for kaf, disc_infos in dolgy.items():
        sheet = wb[kaf]
        for j in range(len(disc_infos)):
            disc = disc_infos[j]
            empty_counter = 0
            print(f"{disc[0]} {disc[1]}")
            res_list = []
            for i in range(1, 100):
                if sheet[f"E{i}"].value is None:
                    empty_counter += 1
                elif f"{disc[0]} {disc[1]}" in sheet[f"E{i}"].value:
                    res_list.append(i)
                elif disc[0] in sheet[f"E{i}"].value:
                    bad_words = ["(Экзамен)", "(Зачет)", "(Курсовая работа)"]
                    for bad_word in bad_words:
                        if bad_word in sheet[f"E{i}"].value:
                            break
                    else:
                        res_list.append(i)
            dolgy[kaf][j].append(res_list)
    return dolgy
                    

def center_el(sheet_el):
    alignment_obj = copy(sheet_el.alignment)
    alignment_obj.horizontal = 'center'
    alignment_obj.vertical = 'center'
    return alignment_obj

def excel_creator(dop_ved_dict, format_dict):
    readme_book = openpyxl.load_workbook(README_FILE)
    Fill = PatternFill(start_color='DCD0FF',
                   end_color='DCD0FF',
                   fill_type='solid')
    
    Fill_malva = PatternFill(start_color='ECE5FF',
                   end_color='ECE5FF',
                   fill_type='solid')
    
    Fill_white = PatternFill(start_color='FFFFFF',
                   end_color='FFFFFF',
                   fill_type='solid')
    
    Thins = Side(border_style="medium", color="8443A6")


    book = openpyxl.Workbook()

    sheet = book.active

    sheet["A1"] = "Кафедра"
    sheet["B1"] = "Дисциплина"
    sheet["C1"] = "Ведомость/допуск"
    sheet["E1"] = "Преподаватель"
    sheet["F1"] = "Предмет"
    sheet["G1"] = "Формат сдачи"
    sheet["H1"] = "БРС"
    sheet["I1"] = "Кабинет"
    sheet["J1"] = "Дата"
    sheet["K1"] = "Время"
    sheet["L1"] = "Примечание"
    sheet["M1"] = "Ссылка 1"
    sheet["N1"] = "Ссылка 2"
    sheet["O1"] = "Ссылка 3"

    sheet.column_dimensions['A'].width = 10
    sheet.column_dimensions['B'].width = 50
    sheet.column_dimensions['C'].width = 20



    i = 1
    color_counter = 0
    for kaf, disc_info in dop_ved_dict.items():
        i += 1
        sheet[f"A{i}"] = kaf
        sheet[f"A{i}"].alignment = center_el(sheet[f"A{i}"])
        sheet[f"A{i}"].fill = Fill
        sheet[f"A{i}"].border = Border(top=Thins, bottom=Thins, left=Thins, right=Thins)

        readme_sheet = readme_book[kaf]
        k = 0
        for disc in disc_info:
            sheet[f"B{i}"] = disc[0] + " " + disc[1]
            sheet[f"B{i}"].font = copy(readme_sheet[f"B{format_dict[kaf][k]}"].font)
            sheet[f"B{i}"].alignment = center_el(readme_sheet[f"B{format_dict[kaf][k]}"])
            sheet[f"B{i}"].fill = Fill
            sheet[f"B{i}"].border = Border(top=Thins, bottom=Thins, left=Thins, right=Thins)
            sheet[f"D{i}"].border = Border(top=Thins, bottom=Thins, left=Thins, right=Thins)

            sheet[f"C{i}"] = disc[2]
            sheet[f"C{i}"].font = copy(readme_sheet[f"C{format_dict[kaf][k]}"].font)
            sheet[f"C{i}"].alignment = center_el(readme_sheet[f"C{format_dict[kaf][k]}"])
            sheet[f"C{i}"].fill = Fill
            sheet[f"C{i}"].border = Border(top=Thins, bottom=Thins, left=Thins, right=Thins) 

            for el in disc[3]:
                sheet.row_dimensions[i].height = 100
                l_str = "EFGHIJKLMNO"
                readme_str = "DEFGHIJKLMN"
                for j in range(len(l_str)):
                    sheet[f"{l_str[j]}{i}"] = readme_sheet[f"{readme_str[j]}{el}"].value
                    sheet[f"{l_str[j]}{i}"].font = copy(readme_sheet[f"{readme_str[j]}{el}"].font)
                    sheet[f"{l_str[j]}{i}"].alignment = copy(readme_sheet[f"{readme_str[j]}{el}"].alignment)
                    sheet[f"{l_str[j]}{i}"].border = Border(top=Thins, bottom=Thins, left=Thins, right=Thins) 
                    if color_counter%2 == 0:
                        sheet[f"{l_str[j]}{i}"].fill = Fill_white
                    else:
                        sheet[f"{l_str[j]}{i}"].fill = Fill_malva

                    sheet[f"{l_str[j]}{i}"].number_format = readme_sheet[f"{readme_str[j]}{el}"].number_format
                color_counter += 1
                i += 1
            k += 1
            i += 1
    
    book.save("results.xlsx")
    book.close()
    readme_book.close()


def printer(dop_ved_dict):
    for kaf, disc_info in dop_ved_dict.items():
        print(kaf) 
        for disc in disc_info:
            print(disc[0] + " " + disc[1])
            print(disc[2])
            for item in disc[3]:
                print(item)
            print("\n")


if __name__ == "__main__":
    data_file = run_sender()
    README_FILE = os.path.join(READFILE_DIR, data_file)
    dolgy = parser()
    dolgy = [f'{i[0]} ({i[1]})' for i in dolgy]
    temp_tup = dop_ved_finder(dolgy)
    dop_ved_dict, format_dict = temp_tup[0], temp_tup[1]
    dop_ved_dict = main_info_finder(dop_ved_dict)
    excel_creator(dop_ved_dict, format_dict)
    